#!/usr/bin/env python3
"""
Download NSW FuelCheck price history files and combine into a SQLite database.

Data source: https://data.nsw.gov.au/data/api/3/action/package_show?id=a97a46fc-2bdd-4b90-ac7f-0cb1e8d7ac3b

File structure variations:
- XLSX (2016 early): header on row 1, full data per row
- XLSX (2019+): title row + blank row + header on row 3, station cols sparse (forward-fill needed)
- CSV: UTF-8 BOM, header on row 1, full data per row
"""

import json
import os
import sqlite3
import sys
import urllib.request
from pathlib import Path

import openpyxl
import pandas as pd

API_URL = "https://data.nsw.gov.au/data/api/3/action/package_show?id=a97a46fc-2bdd-4b90-ac7f-0cb1e8d7ac3b"
RAW_DIR = Path("raw")
DB_PATH = Path("fuelhistory.db")

EXPECTED_COLS = ["ServiceStationName", "Address", "Suburb", "Postcode", "Brand", "FuelCode", "PriceUpdatedDate", "Price"]
STATION_COLS = ["ServiceStationName", "Address", "Suburb", "Postcode", "Brand"]


def fetch_resources():
    with urllib.request.urlopen(API_URL) as r:
        data = json.loads(r.read())
    resources = data["result"]["resources"]
    price_history = []
    for res in resources:
        name = res.get("name", "")
        fmt = res.get("format", "").lower()
        url = res.get("url", "")
        # Skip non-data resources
        if not url.endswith((".xlsx", ".csv")):
            continue
        if "price" not in name.lower() and "price" not in url.lower():
            continue
        price_history.append({"name": name, "format": fmt, "url": url, "id": res.get("id", "")})
    return price_history


def download_file(url: str, dest: Path) -> bool:
    if dest.exists():
        print(f"  [skip] already downloaded: {dest.name}")
        return True
    try:
        print(f"  [dl] {dest.name}")
        urllib.request.urlretrieve(url, dest)
        return True
    except Exception as e:
        print(f"  [err] failed to download {url}: {e}")
        return False


def read_xlsx(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return pd.DataFrame(columns=EXPECTED_COLS)

    # Find the header row: look for row containing 'ServiceStationName'
    header_idx = None
    for i, row in enumerate(rows[:10]):
        row_strs = [str(c).strip() if c is not None else "" for c in row]
        if "ServiceStationName" in row_strs:
            header_idx = i
            break

    if header_idx is None:
        print(f"    [warn] could not find header in {path.name}, skipping")
        return pd.DataFrame(columns=EXPECTED_COLS)

    headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[header_idx])]
    data_rows = rows[header_idx + 1:]

    df = pd.DataFrame(data_rows, columns=headers)

    # August 2016 uses 'FuelType' instead of 'FuelCode'
    if "FuelType" in df.columns and "FuelCode" not in df.columns:
        df = df.rename(columns={"FuelType": "FuelCode"})

    # Forward-fill station columns (sparse rows in later files)
    for col in STATION_COLS:
        if col in df.columns:
            df[col] = df[col].replace("", None).ffill()

    # Drop rows with no FuelCode or Price (blank separators etc.)
    df = df.dropna(subset=["FuelCode", "Price"])

    return df[EXPECTED_COLS] if all(c in df.columns for c in EXPECTED_COLS) else df


def read_csv(path: Path) -> pd.DataFrame:
    try:
        df = pd.read_csv(path, encoding="utf-8-sig", dtype=str)
    except Exception as e:
        print(f"    [warn] csv read error {path.name}: {e}")
        return pd.DataFrame(columns=EXPECTED_COLS)

    df.columns = [c.strip() for c in df.columns]
    df = df.dropna(subset=["FuelCode", "Price"])
    return df[EXPECTED_COLS] if all(c in df.columns for c in EXPECTED_COLS) else df


def normalise(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Normalise Price to float
    df["Price"] = pd.to_numeric(df["Price"], errors="coerce")

    # Normalise Postcode to string (sometimes comes as int)
    df["Postcode"] = df["Postcode"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

    # Normalise PriceUpdatedDate to ISO datetime string.
    # Files use two formats:
    #   - ISO:   "2024-07-01 00:04:36"  (CSV files, YYYY-MM-DD)
    #   - D/M/Y: "1/10/2019 12:02:26 AM"  (XLSX files, DD/MM/YYYY with AM/PM)
    # Detect format by checking if first non-null value starts with 4 digits.
    sample = df["PriceUpdatedDate"].dropna().iloc[0] if not df["PriceUpdatedDate"].dropna().empty else ""
    if isinstance(sample, str) and len(sample) >= 4 and sample[:4].isdigit() and sample[4] == "-":
        df["PriceUpdatedDate"] = pd.to_datetime(df["PriceUpdatedDate"], format="ISO8601", errors="coerce")
    else:
        df["PriceUpdatedDate"] = pd.to_datetime(df["PriceUpdatedDate"], dayfirst=True, errors="coerce")
    df["PriceUpdatedDate"] = df["PriceUpdatedDate"].dt.strftime("%Y-%m-%d %H:%M:%S")

    # Strip whitespace from string columns
    for col in ["ServiceStationName", "Address", "Suburb", "Brand", "FuelCode"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    return df


def main():
    RAW_DIR.mkdir(exist_ok=True)

    print("Fetching resource list...")
    resources = fetch_resources()
    print(f"Found {len(resources)} price history files")

    DB_PATH.unlink(missing_ok=True)
    con = sqlite3.connect(DB_PATH)
    errors = []
    total_rows = 0
    files_written = 0

    for res in resources:
        url = res["url"]
        ext = ".csv" if url.lower().endswith(".csv") else ".xlsx"
        filename = res["id"] + ext
        dest = RAW_DIR / filename

        print(f"\n{res['name']}")
        if not download_file(url, dest):
            errors.append(res["name"])
            continue

        try:
            if ext == ".xlsx":
                df = read_xlsx(dest)
            else:
                df = read_csv(dest)

            if df.empty:
                print(f"    [warn] empty dataframe for {dest.name}")
                continue

            df = normalise(df)
            df["source_file"] = res["name"]

            # Write incrementally — first file creates table, rest append
            if_exists = "replace" if files_written == 0 else "append"
            df.to_sql("prices", con, index=False, if_exists=if_exists)
            con.commit()

            total_rows += len(df)
            files_written += 1
            print(f"    {len(df):,} rows  (running total: {total_rows:,})")
        except Exception as e:
            print(f"    [err] {dest.name}: {e}")
            errors.append(res["name"])

    if files_written == 0:
        print("No data loaded, exiting.")
        con.close()
        sys.exit(1)

    print(f"\nBuilding indexes...")
    con.execute("CREATE INDEX IF NOT EXISTS idx_date ON prices(PriceUpdatedDate)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_fuel ON prices(FuelCode)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_station ON prices(ServiceStationName)")
    con.commit()

    row_count = con.execute("SELECT COUNT(*) FROM prices").fetchone()[0]
    con.close()

    print(f"Done. {row_count:,} rows in {DB_PATH}")

    if errors:
        print(f"\nFailed files ({len(errors)}):")
        for e in errors:
            print(f"  - {e}")


if __name__ == "__main__":
    main()
